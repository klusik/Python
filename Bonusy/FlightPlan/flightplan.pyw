#!/usr/bin/env python3
"""
Simple Navigraph route to Excel flight log generator.

This script provides a Tkinter GUI where the user pastes a route such as

    KSAN DCT MZB DCT OCN DCT SLI DCT VTU DCT RZS DCT MQO DCT BSR DCT PYE DCT ENI DCT FOT DCT KEKA

The script parses the route into legs and generates an Excel flight log
containing only the columns needed for in flight pencil note keeping:

    Leg, From, To, Planned Altitude, Assigned Altitude, Wind,
    GS, Time, Fuel Used, Fuel Remaining, FF, Remarks.
"""

import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    Workbook = None
else:
    pass


class FlightLeg:
    """
    Data container representing a single leg in the route.

    Attributes
    ----------
    index : int
        Leg index, starting from 1 for the first leg.
    origin : str
        Identifier of the waypoint or airport at the start of the leg.
    destination : str
        Identifier of the waypoint or airport at the end of the leg.
    """

    def __init__(self, index: int, origin: str, destination: str) -> None:
        """
        Initialize a new FlightLeg instance.

        Parameters
        ----------
        index : int
            Sequential leg index, starting from 1.
        origin : str
            Identifier of the origin waypoint or airport.
        destination : str
            Identifier of the destination waypoint or airport.
        """
        self.index = index
        self.origin = origin
        self.destination = destination

    def as_tuple(self) -> tuple[int, str, str]:
        """
        Return the leg as a tuple.

        Returns
        -------
        tuple[int, str, str]
            Tuple containing (index, origin, destination).
        """
        return self.index, self.origin, self.destination


class FlightLog:
    """
    Parse a Navigraph style route string and build an Excel flight log.

    The class is independent from the GUI and can be used directly from
    Python code if needed. It performs three main steps:

    1. Parse the route string into a list of waypoint tokens.
    2. Convert the tokens into a sequence of FlightLeg objects.
    3. Build and save an openpyxl Workbook using a fixed flight log layout.
    """

    def __init__(self, route: str, treat_ends_as_airports: bool = True) -> None:
        """
        Construct a FlightLog instance.

        Parameters
        ----------
        route : str
            Route string in Navigraph style, for example
            "KSAN DCT MZB DCT OCN DCT KEKA".
        treat_ends_as_airports : bool, optional
            Flag kept for future extension. Currently the first token is
            used as departure and the last token as destination either way,
            but this option allows changing the behavior later without
            touching the GUI.
        """
        self.raw_route = route.strip()
        self.treat_ends_as_airports = treat_ends_as_airports

        self.tokens = self._parse_route_tokens(self.raw_route)
        if len(self.tokens) < 2:
            raise ValueError("Route must contain at least departure and destination.")

        self.departure = self.tokens[0]
        self.destination = self.tokens[-1]

        self.legs = self._build_legs(self.tokens)

    @staticmethod
    def _parse_route_tokens(route: str) -> list[str]:
        """
        Parse the raw route string into a list of waypoint tokens.

        All identifiers are converted to upper case and passive keywords
        such as "DCT" or "DIRECT" are removed.

        Parameters
        ----------
        route : str
            Raw route string.

        Returns
        -------
        list[str]
            Ordered list of waypoints and airports.
        """
        raw_tokens = route.upper().split()
        skip_tokens = {"DCT", "DIRECT"}
        return [token for token in raw_tokens if token not in skip_tokens]

    @staticmethod
    def _build_legs(tokens: list[str]) -> list[FlightLeg]:
        """
        Build a list of FlightLeg objects from ordered waypoint tokens.

        Parameters
        ----------
        tokens : list[str]
            Ordered list of waypoints and airports.

        Returns
        -------
        list[FlightLeg]
            List of legs, where leg i connects tokens[i] to tokens[i + 1].
        """
        legs: list[FlightLeg] = []
        for idx in range(len(tokens) - 1):
            legs.append(FlightLeg(index=idx + 1, origin=tokens[idx], destination=tokens[idx + 1]))
        return legs

    def _create_workbook(self) -> "Workbook":
        """
        Create and populate an openpyxl Workbook with the flight log layout.

        The workbook contains a single sheet with the following structure:

        - Row 1: Title "Flight Log".
        - Row 2: Route line.
        - Row 3: Departure and destination labels.
        - Row 5: Column headers.
        - Row 6 onward: One row per leg, with only Leg, From and To prefilled.

        Returns
        -------
        Workbook
            A fully populated openpyxl Workbook instance.

        Raises
        ------
        RuntimeError
            If openpyxl is not available in the current Python environment.
        """
        if Workbook is None:
            raise RuntimeError("openpyxl is not installed. Install it with 'pip install openpyxl'.")

        wb = Workbook()
        ws = wb.active
        ws.title = "Flight Log"

        title_font = Font(size=14, bold=True)
        header_font = Font(size=11, bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title and route summary
        ws["A1"] = "Flight Log"
        ws["A1"].font = title_font
        ws["A1"].alignment = left_align

        ws["A2"] = "Route:"
        ws["B2"] = self.raw_route
        ws["A2"].font = header_font
        ws["A2"].alignment = left_align

        ws["A3"] = "Departure:"
        ws["B3"] = self.departure
        ws["C3"] = "Destination:"
        ws["D3"] = self.destination
        ws["A3"].font = header_font
        ws["C3"].font = header_font
        ws["A3"].alignment = left_align
        ws["C3"].alignment = left_align

        # Column headers
        start_row = 5
        headers = [
            "Leg",
            "From",
            "To",
            "Planned Altitude",
            "Assigned Altitude",
            "Wind",
            "GS",
            "Time",
            "Fuel Used",
            "Fuel Remaining",
            "FF",
            "Remarks",
        ]

        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_index, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Data rows
        data_start_row = start_row + 1
        for leg_offset, leg in enumerate(self.legs):
            row = data_start_row + leg_offset
            leg_index, origin, destination = leg.as_tuple()
            values: list[object] = [leg_index, origin, destination]

            # Extend with empty strings for remaining columns
            while len(values) < len(headers):
                values.append("")

            for col_index, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_index, value=value)
                cell.border = thin_border
                if col_index <= 3:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        # Column widths (first column widened so the header is fully visible)
        column_widths = {
            1: 20,  # Leg (wider to avoid truncation)
            2: 12,  # From
            3: 12,  # To
            4: 18,  # Planned Altitude
            5: 18,  # Assigned Altitude
            6: 10,  # Wind
            7: 10,  # GS
            8: 12,  # Time
            9: 12,  # Fuel Used
            10: 15, # Fuel Remaining
            11: 8,  # FF
            12: 20, # Remarks
        }

        for col_idx, width in column_widths.items():
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = width

        # Freeze header row
        ws.freeze_panes = ws["A6"]

        return wb

    def generate_excel(self, output_path: str) -> None:
        """
        Generate the Excel file on disk.

        Parameters
        ----------
        output_path : str
            Path where the .xlsx file should be written.

        Raises
        ------
        IOError
            If the workbook cannot be saved to the specified location.
        """
        workbook = self._create_workbook()
        try:
            workbook.save(output_path)
        except OSError as exc:
            raise IOError(f"Failed to write Excel file: {exc}") from exc


class FlightLogGUI:
    """
    Tkinter based graphical interface for the FlightLog generator.

    The GUI allows the user to:
    - Enter a Navigraph style route string.
    - Optionally specify that first and last tokens are airports.
    - Choose an output filename.
    - Generate the Excel flight log with a single button press.
    """

    def __init__(self, master: tk.Tk) -> None:
        """
        Initialize the GUI and build all widgets.

        Parameters
        ----------
        master : tk.Tk
            Tkinter root window instance.
        """
        self.master = master
        self.master.title("Navigraph Route to Excel Flight Log")

        self.route_var = tk.StringVar()
        self.airports_var = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="")

        self._build_layout()

    def _build_layout(self) -> None:
        """
        Create and arrange widgets inside the main window.
        """
        padding = {"padx": 10, "pady": 5}

        ttk.Label(self.master, text="Route (Navigraph format):").grid(
            row=0, column=0, sticky="w", **padding
        )

        ttk.Entry(self.master, textvariable=self.route_var, width=80).grid(
            row=1, column=0, columnspan=3, sticky="we", **padding
        )

        ttk.Checkbutton(
            self.master,
            text="Treat first and last as airports",
            variable=self.airports_var,
        ).grid(row=2, column=0, columnspan=3, sticky="w", **padding)

        ttk.Button(
            self.master,
            text="Generate Excel Flight Log",
            command=self.on_generate_clicked,
        ).grid(row=3, column=0, sticky="w", **padding)

        ttk.Label(self.master, textvariable=self.status_var, foreground="blue").grid(
            row=4, column=0, columnspan=3, sticky="w", **padding
        )

        self.master.columnconfigure(0, weight=1)

    def on_generate_clicked(self) -> None:
        """
        Callback for the "Generate Excel Flight Log" button.

        Validates the input route, prompts the user for an output path,
        and then delegates Excel generation to the FlightLog class.
        """
        route = self.route_var.get().strip()
        if not route:
            messagebox.showerror("Input error", "Please enter a route string.")
            return

        if Workbook is None:
            messagebox.showerror(
                "Dependency missing",
                "The library 'openpyxl' is not installed.\nInstall it with:\n\npip install openpyxl",
            )
            return

        default_filename = "flight_log.xlsx"
        filetypes = [("Excel Workbook", "*.xlsx"), ("All Files", "*.*")]
        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=filetypes,
            title="Save Flight Log As",
        )

        if not output_path:
            return

        try:
            flight_log = FlightLog(
                route=route,
                treat_ends_as_airports=self.airports_var.get(),
            )
            flight_log.generate_excel(output_path)
        except Exception as exc:
            messagebox.showerror("Error", f"Failed to generate Excel file:\n{exc}")
            return

        self.status_var.set(f"Flight log saved to: {output_path}")
        messagebox.showinfo("Success", f"Flight log successfully saved to:\n{output_path}")


def main() -> None:
    """
    Application entry point.

    This function validates that openpyxl is available, creates the
    Tkinter root window, constructs the GUI, and enters the Tkinter
    main loop.
    """
    if Workbook is None:
        msg = (
            "The library 'openpyxl' is required but is not installed.\n"
            "Install it with:\n\n"
            "    pip install openpyxl\n"
        )
        sys.stderr.write(msg)
        try:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Dependency missing", msg)
            root.destroy()
        except Exception:
            pass
        sys.exit(1)

    root = tk.Tk()
    FlightLogGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
